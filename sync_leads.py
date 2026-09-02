#!/usr/bin/env python3
"""
Harmonium Leads Auto-Sync para GitHub Actions
Sincroniza emails de no-reply@harmonium.design al Excel en Google Drive
"""

import re
import json
import os
import pickle
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import base64


class HarmoniumLeadsSync:
    GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.readonly', 
                    'https://www.googleapis.com/auth/gmail.modify']
    DRIVE_SCOPES = ['https://www.googleapis.com/auth/drive']
    
    def __init__(self):
        self.gmail_service = self._authenticate_gmail()
        self.drive_service = self._authenticate_drive()
        self.processed_emails = set()
        self.load_processed_emails()
    
    def _authenticate_gmail(self):
        """Autentica con Gmail API usando token.pickle"""
        creds = None
        
        # Intenta cargar token.pickle desde GitHub Actions
        token_b64 = os.getenv('GOOGLE_TOKEN_PICKLE')
        if token_b64:
            try:
                token_data = base64.b64decode(token_b64)
                creds = pickle.loads(token_data)
            except Exception as e:
                print(f"Error decodificando token: {e}")
                return None
        
        # Si no hay token en env, intenta archivo local
        if not creds and Path('token.pickle').exists():
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)
        
        # Refresca el token si es necesario
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        
        if not creds:
            raise ValueError("No credentials available")
        
        return build('gmail', 'v1', credentials=creds)
    
    def _authenticate_drive(self):
        """Autentica con Google Drive API usando el mismo token"""
        creds = None
        
        token_b64 = os.getenv('GOOGLE_TOKEN_PICKLE')
        if token_b64:
            try:
                token_data = base64.b64decode(token_b64)
                creds = pickle.loads(token_data)
            except Exception as e:
                print(f"Error decodificando token: {e}")
                return None
        
        if not creds and Path('token.pickle').exists():
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)
        
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        
        if not creds:
            raise ValueError("No credentials available")
        
        return build('drive', 'v3', credentials=creds)
    
    def load_processed_emails(self):
        """Carga los IDs de emails ya procesados"""
        log_file = Path('.processed_emails.json')
        if log_file.exists():
            with open(log_file, 'r') as f:
                data = json.load(f)
                self.processed_emails = set(data.get('ids', []))
    
    def save_processed_emails(self):
        """Guarda los IDs de emails procesados"""
        with open('.processed_emails.json', 'w') as f:
            json.dump({'ids': list(self.processed_emails)}, f)
    
    def fetch_new_emails(self):
        """Obtiene emails sin leer de no-reply@harmonium.design"""
        try:
            results = self.gmail_service.users().messages().list(
                userId='me',
                q='from:no-reply@harmonium.design is:unread',
                maxResults=50
            ).execute()
            
            messages = results.get('messages', [])
            return messages
        except Exception as e:
            print(f"Error al buscar emails: {e}")
            return []
    
    def extract_lead_data(self, message_id):
        """Extrae datos del lead del email"""
        try:
            msg = self.gmail_service.users().messages().get(
                userId='me', id=message_id, format='full').execute()
            
            headers = msg['payload']['headers']
            date_str = next(h['value'] for h in headers if h['name'] == 'Date')
            
            body_data = msg['payload'].get('parts', [{}])[0].get('body', {})
            body_text = body_data.get('data', '')
            if not body_text and 'parts' in msg['payload']:
                for part in msg['payload']['parts']:
                    if 'body' in part:
                        body_text = part['body'].get('data', '')
                        break
            
            if body_text:
                body_text = base64.urlsafe_b64decode(body_text).decode('utf-8')
            
            lead = self._parse_lead_text(body_text)
            lead['fecha'] = self._parse_date(date_str)
            lead['email_id'] = message_id
            
            return lead
        except Exception as e:
            print(f"Error extrayendo datos de {message_id}: {e}")
            return None
    
    def _parse_lead_text(self, text):
        """Parsea el cuerpo del email"""
        lead = {
            'nombre': '',
            'email': '',
            'telefono': '',
            'tipo_cliente': '',
            'ubicacion': '',
            'mensaje': ''
        }
        
        patterns = {
            'nombre': r'Nombre:\s*(.+?)(?:\n|$)',
            'email': r'Email:\s*(.+?)(?:\n|$)',
            'telefono': r'Teléfono:\s*(.+?)(?:\n|$)',
            'tipo_cliente': r'Tipo de cliente:\s*(.+?)(?:\n|$)',
            'ubicacion': r'Ubicación:\s*(.+?)(?:\n|$)',
            'mensaje': r'Mensaje:\s*(.+?)(?:\n\n|$)'
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
            if match:
                lead[key] = match.group(1).strip()
        
        return lead
    
    def _parse_date(self, date_str):
        """Convierte fecha RFC 2822"""
        try:
            from email.utils import parsedate_to_datetime
            dt = parsedate_to_datetime(date_str)
            return dt.strftime('%Y-%m-%d')
        except:
            return datetime.now().strftime('%Y-%m-%d')
    
    def find_file_in_drive(self, filename):
        """Busca un archivo en Google Drive"""
        try:
            results = self.drive_service.files().list(
                q=f"name='{filename}' and trashed=false",
                spaces='drive',
                fields='files(id, name)',
                pageSize=1
            ).execute()
            
            files = results.get('files', [])
            return files[0]['id'] if files else None
        except Exception as e:
            print(f"Error buscando archivo: {e}")
            return None
    
    def download_excel_from_drive(self, file_id, local_path):
        """Descarga el Excel desde Google Drive"""
        try:
            request = self.drive_service.files().get_media(fileId=file_id)
            with open(local_path, 'wb') as f:
                f.write(request.execute())
            print(f"✓ Excel descargado de Drive")
        except Exception as e:
            print(f"Error descargando: {e}")
    
    def upload_excel_to_drive(self, file_id, local_path):
        """Sube el Excel a Google Drive"""
        try:
            from googleapiclient.http import MediaFileUpload
            media = MediaFileUpload(local_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.drive_service.files().update(fileId=file_id, media_body=media).execute()
            print(f"✓ Excel actualizado en Drive")
        except Exception as e:
            print(f"Error subiendo: {e}")
    
    def update_excel(self, leads, excel_path):
        """Agrega leads al Excel"""
        if not leads:
            return 0
        
        if not Path(excel_path).exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Leads"
            
            headers = ['Fecha lead', 'Nombre', 'Email', 'Teléfono', 
                      'Tipo de cliente', 'Ubicación', 'Interés / Producto', 
                      'Mensaje', 'URL página', 'Estado', 'Fecha registro', 
                      'ID Gmail', 'GCLID', 'Estado comercial', 
                      'Fecha conversión', 'Valor €']
            
            ws.append(headers)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            wb.save(excel_path)
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        count = 0
        for lead in leads:
            if lead['email_id'] in self.processed_emails:
                continue
            
            row = [
                lead.get('fecha'),
                lead.get('nombre'),
                lead.get('email'),
                lead.get('telefono'),
                lead.get('tipo_cliente'),
                lead.get('ubicacion'),
                lead.get('mensaje'),
                '', '', '', '',
                lead.get('email_id'),
                '', 'Nuevo', '', ''
            ]
            
            ws.append(row)
            self.processed_emails.add(lead['email_id'])
            count += 1
        
        if count > 0:
            wb.save(excel_path)
            print(f"✓ {count} leads agregados")
        
        return count
    
    def mark_emails_read(self, message_ids):
        """Marca emails como leídos"""
        try:
            for msg_id in message_ids:
                self.gmail_service.users().messages().modify(
                    userId='me',
                    id=msg_id,
                    body={'removeLabelIds': ['UNREAD']}
                ).execute()
        except Exception as e:
            print(f"Error marcando como leídos: {e}")
    
    def sync(self):
        """Sincronización completa"""
        print(f"\n🔄 Sincronización: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        excel_filename = 'Leads_Harmonium_marcado_amarillo.xlsx'
        local_excel = excel_filename
        
        file_id = self.find_file_in_drive(excel_filename)
        if file_id:
            self.download_excel_from_drive(file_id, local_excel)
        
        messages = self.fetch_new_emails()
        if not messages:
            print("✓ Sin emails nuevos")
            return 0
        
        print(f"📧 {len(messages)} emails encontrados")
        
        leads = []
        msg_ids = []
        for msg in messages:
            lead = self.extract_lead_data(msg['id'])
            if lead:
                leads.append(lead)
                msg_ids.append(msg['id'])
        
        count = self.update_excel(leads, local_excel)
        
        if count > 0 and file_id:
            self.upload_excel_to_drive(file_id, local_excel)
            self.mark_emails_read(msg_ids)
            self.save_processed_emails()
        
        return count


if __name__ == "__main__":
    sync = HarmoniumLeadsSync()
    sync.sync()
